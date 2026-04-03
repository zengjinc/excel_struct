import openpyxl
import json
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List
from config_parser import ExcelConfig, SheetConfig, ErlFun, LuaFun, PreFormat, Field


class ExcelGenerator:
    def __init__(self):
        self.wb = None

    def generate(self, config: ExcelConfig, output_path: str = None) -> openpyxl.Workbook:
        self.wb = openpyxl.Workbook()

        for sheet in config.sheets:
            self._generate_sheet(sheet)

        if output_path is None:
            output_path = config.output_file

        self.wb.save(output_path)
        return self.wb

    def _generate_sheet(self, sheet: SheetConfig):
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        ws = self.wb.create_sheet(title=sheet.name)
        current_row = 1

        current_row = self._write_erl_name(ws, sheet, current_row)
        current_row = self._write_erl_include(ws, sheet, current_row)
        current_row = self._write_erl_funs(ws, sheet, current_row)
        current_row = self._write_lua_name(ws, sheet, current_row)
        current_row = self._write_lua_funs(ws, sheet, current_row)
        current_row = self._write_pre_format(ws, sheet, current_row)
        current_row = self._write_fields(ws, sheet, current_row)
        current_row = self._write_notes(ws, sheet, current_row)
        current_row = self._write_values(ws, sheet, current_row)

        self._apply_styles(ws, sheet)

    def _write_erl_name(self, ws, sheet: SheetConfig, row: int) -> int:
        ws.cell(row=row, column=1, value='ERL_NAME')
        ws.cell(row=row, column=2, value=sheet.erl_name)
        return row + 1

    def _write_erl_include(self, ws, sheet: SheetConfig, row: int) -> int:
        if not sheet.erl_include:
            return row

        ws.cell(row=row, column=1, value='ERL_INCLUDE')
        for idx, include in enumerate(sheet.erl_include):
            ws.cell(row=row, column=2 + idx, value=include)
        return row + 1

    def _write_erl_funs(self, ws, sheet: SheetConfig, row: int) -> int:
        if not sheet.erl_funs:
            return row

        for erl_fun in sheet.erl_funs:
            ws.cell(row=row, column=1, value='ERL_FUN')
            ws.cell(row=row, column=2, value=erl_fun.name)

            params_json = self._dict_to_json(erl_fun.params)
            ws.cell(row=row, column=3, value=params_json)

            return_json = self._dict_to_json({"return": erl_fun.return_type})
            ws.cell(row=row, column=4, value=return_json)

            when_json = self._dict_to_json({"when": erl_fun.when_clause})
            ws.cell(row=row, column=5, value=when_json)

            note_json = self._dict_to_json({"note": erl_fun.note})
            ws.cell(row=row, column=6, value=note_json)

            if erl_fun.filter_config:
                filter_json = self._dict_to_json({"filter": erl_fun.filter_config})
                ws.cell(row=row, column=7, value=filter_json)

            row += 1

        return row

    def _write_lua_name(self, ws, sheet: SheetConfig, row: int) -> int:
        ws.cell(row=row, column=1, value='LUA_NAME')
        ws.cell(row=row, column=2, value=sheet.lua_name)
        return row + 1

    def _write_lua_funs(self, ws, sheet: SheetConfig, row: int) -> int:
        if not sheet.lua_funs:
            return row

        for lua_fun in sheet.lua_funs:
            ws.cell(row=row, column=1, value='LUA_FUN')
            ws.cell(row=row, column=2, value=lua_fun.name)

            params_json = self._dict_to_json(lua_fun.params)
            ws.cell(row=row, column=3, value=params_json)

            if lua_fun.default:
                default_json = self._dict_to_json(lua_fun.default)
                ws.cell(row=row, column=4, value=default_json)

            if lua_fun.return_type:
                return_json = self._dict_to_json({"return": lua_fun.return_type})
                ws.cell(row=row, column=5, value=return_json)

            if lua_fun.sub_key:
                sub_key_json = self._dict_to_json({"sub_key": lua_fun.sub_key})
                ws.cell(row=row, column=6, value=sub_key_json)

            if lua_fun.key_split:
                key_split_json = self._dict_to_json({"key_split": lua_fun.key_split})
                ws.cell(row=row, column=7, value=key_split_json)

            if lua_fun.filter_config:
                filter_json = self._dict_to_json({"filter": lua_fun.filter_config})
                ws.cell(row=row, column=8, value=filter_json)

            row += 1

        return row

    def _write_pre_format(self, ws, sheet: SheetConfig, row: int) -> int:
        if not sheet.pre_format:
            return row

        ws.cell(row=row, column=1, value='PRE_FORMAT')

        pre_format_json = {
            "key": sheet.pre_format.key,
            "value": sheet.pre_format.value
        }
        if sheet.pre_format.server:
            pre_format_json["server"] = sheet.pre_format.server
        if sheet.pre_format.client:
            pre_format_json["client"] = sheet.pre_format.client

        ws.cell(row=row, column=2, value=self._dict_to_json(pre_format_json))
        return row + 1

    def _write_fields(self, ws, sheet: SheetConfig, row: int) -> int:
        ws.cell(row=row, column=1, value='FIELD')
        for idx, field in enumerate(sheet.fields):
            ws.cell(row=row, column=2 + idx, value=field.name)
        return row + 1

    def _write_notes(self, ws, sheet: SheetConfig, row: int) -> int:
        ws.cell(row=row, column=1, value='NOTE')
        for idx, field in enumerate(sheet.fields):
            ws.cell(row=row, column=2 + idx, value=field.note)
        return row + 1

    def _write_values(self, ws, sheet: SheetConfig, row: int) -> int:
        for value_row in sheet.values:
            ws.cell(row=row, column=1, value='VALUE')
            for idx, value in enumerate(value_row):
                ws.cell(row=row, column=2 + idx, value=value)
            row += 1
        return row

    def _dict_to_json(self, data: dict) -> str:
        return json.dumps(data, ensure_ascii=False)

    def _apply_styles(self, ws, sheet: SheetConfig):
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value in ['ERL_NAME', 'ERL_INCLUDE', 'ERL_FUN', 'LUA_NAME', 'LUA_FUN',
                            'PRE_FORMAT', 'FIELD', 'NOTE', 'VALUE']:
                cell.fill = header_fill
                cell.font = header_font

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
