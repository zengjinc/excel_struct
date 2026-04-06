import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def generate_excel(config, target_path):
    """
    根据配置生成 Excel 文件
    :param config: 配置字典
    :param target_path: 目标文件路径
    """
    # 检查结构是否有差异
    if not _check_structure_diff(config, target_path):
        print(f"Excel 文件结构无差异，跳过生成: {target_path}")
        return

    # 检查目标文件是否存在
    if os.path.exists(target_path):
        # 读取现有文件
        wb = openpyxl.load_workbook(target_path)
        # 保留 VALUE 行数据
        value_data = _extract_value_data(wb)
        # 删除默认的 Sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    else:
        # 创建新文件
        wb = openpyxl.Workbook()
        # 删除默认的 Sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        value_data = {}

    # 处理每个 sheet
    for sheet_config in config['sheets']:
        sheet_name = sheet_config['name']
        sheet_data = sheet_config['config']

        # 检查 sheet 是否存在
        if sheet_name in wb.sheetnames:
            # 保留现有 sheet
            ws = wb[sheet_name]
            # 清除现有的配置信息（保留 VALUE 行数据）
            # 找到 FIELD 行
            field_row = None
            for row in range(1, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == 'FIELD':
                    field_row = row
                    break
            # 清除 FIELD 行之前的所有内容
            if field_row:
                for row in range(1, field_row):
                    for column in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=column).value = None
                # 清除 FIELD 行和 NOTE 行
                note_row = field_row + 1
                value_row = note_row + 1
                for row in range(field_row, value_row):
                    for column in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=column).value = None
            else:
                # 如果没有找到 FIELD 行，清除所有内容
                for row in range(1, ws.max_row + 1):
                    for column in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=column).value = None
            # 写入新的配置信息
            _write_sheet_config(ws, sheet_data)
            # 写入 VALUE 行数据
            if sheet_name in value_data:
                _write_value_data(ws, value_data[sheet_name])
        else:
            # 创建新 sheet
            ws = wb.create_sheet(title=sheet_name)
            # 写入配置信息
            _write_sheet_config(ws, sheet_data)
            # 写入 VALUE 行数据
            if sheet_name in value_data:
                _write_value_data(ws, value_data[sheet_name])

    # 保存文件
    wb.save(target_path)
    print(f"生成 Excel 文件: {target_path}")

def _extract_value_data(wb):
    """
    提取 VALUE 行数据
    :param wb: Workbook 对象
    :return: VALUE 行数据字典，格式为 {sheet_name: {field_name: [values], 'value_rows': [row_numbers], 'blank_lines': [row_numbers]}}
    """
    value_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_value_data = {
            'value_rows': [],
            'blank_lines': []
        }

        # 查找 VALUE 行和空白行
        field_row = None

        # 先找到 FIELD 行和所有 VALUE 行
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == 'FIELD':
                field_row = row
            elif cell_value == 'VALUE':
                sheet_value_data['value_rows'].append(row)
            elif cell_value is None:
                # 检查是否是空白行（所有列都为空）
                is_blank = True
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=row, column=col).value is not None:
                        is_blank = False
                        break
                if is_blank:
                    sheet_value_data['blank_lines'].append(row)

        # 提取字段名称
        fields = []
        if field_row:
            column = 2
            while True:
                field_value = ws.cell(row=field_row, column=column).value
                if field_value:
                    fields.append(field_value)
                    column += 1
                else:
                    break

        # 提取 VALUE 行数据
        if sheet_value_data['value_rows'] and fields:
            # 初始化每个字段的数据列表
            for field in fields:
                sheet_value_data[field] = []

            # 只提取实际的 VALUE 行数据
            for row in sheet_value_data['value_rows']:
                for i, field in enumerate(fields):
                    value = ws.cell(row=row, column=i+2).value
                    sheet_value_data[field].append(value)

        value_data[sheet_name] = sheet_value_data

    return value_data

def _write_sheet_config(ws, sheet_data):
    """
    写入 sheet 配置信息
    :param ws: Worksheet 对象
    :param sheet_data: sheet 配置数据
    """
    # 定义颜色填充
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 颜色代码4
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 颜色代码5

    # 写入 ERL_NAME
    ws.cell(row=1, column=1, value='ERL_NAME').fill = yellow_fill
    ws.cell(row=1, column=2, value=sheet_data.get('ERL_NAME', ''))

    # 写入 ERL_INCLUDE
    ws.cell(row=2, column=1, value='ERL_INCLUDE').fill = yellow_fill
    erl_includes = sheet_data.get('ERL_INCLUDE', [])
    for i, include in enumerate(erl_includes, 2):
        ws.cell(row=2, column=i, value=include)

    # 写入 ERL_FUN
    erl_ffuns = sheet_data.get('ERL_FFUN', [])
    for i, ffun in enumerate(erl_ffuns, 3):
        ws.cell(row=i, column=1, value='ERL_FUN').fill = yellow_fill
        ws.cell(row=i, column=2, value=ffun.get('name', ''))
        # 写入 params
        params = ffun.get('params', {})
        key = params.get('key', [])
        # 如果key列表全为空字符串，替换为[]
        if isinstance(key, list) and all(item == '' for item in key):
            key = []
        value = params.get('value', [])
        params_str = '{"key":' + str(key) + ', "value":' + str(value) + '}'
        ws.cell(row=i, column=3, value=params_str)
        # 写入 return (only if not empty)
        return_value = ffun.get('return', '')
        if return_value:
            return_str = '{"return":"' + return_value + '"}'
            ws.cell(row=i, column=4, value=return_str)
        # 写入 when (only if not empty)
        when_value = ffun.get('when', '')
        if when_value:
            when_str = '{"when":"' + when_value + '"}'
            ws.cell(row=i, column=5, value=when_str)
        # 写入 note (only if not empty)
        note_value = ffun.get('note', '')
        if note_value:
            note_str = '{"note":"' + note_value + '"}'
            ws.cell(row=i, column=6, value=note_str)
        # 写入 fun_note (only if not empty)
        fun_note_value = ffun.get('fun_note', [])
        if fun_note_value:
            fun_note_str = '{"fun_note":' + str(fun_note_value) + '}'
            ws.cell(row=i, column=7, value=fun_note_str)

    # 写入 LUA_NAME
    lua_name_row = 3 + len(erl_ffuns)
    ws.cell(row=lua_name_row, column=1, value='LUA_NAME').fill = yellow_fill
    ws.cell(row=lua_name_row, column=2, value=sheet_data.get('LUA_NAME', ''))

    # 写入 LUA_FUN
    lua_funs = sheet_data.get('LUA_FUN', [])
    for i, fun in enumerate(lua_funs, lua_name_row + 1):
        ws.cell(row=i, column=1, value='LUA_FUN').fill = yellow_fill
        ws.cell(row=i, column=2, value=fun.get('name', ''))
        # 写入 params
        params = fun.get('params', {})
        key = params.get('key', [])
        # 如果key列表全为空字符串，替换为[]
        if isinstance(key, list) and all(item == '' for item in key):
            key = []
        value = params.get('value', [])
        params_str = '{"key":' + str(key) + ', "value":' + str(value) + '}'
        ws.cell(row=i, column=3, value=params_str)
        # 写入 return (only if not empty)
        return_value = fun.get('return', '')
        if return_value:
            return_str = '{"return":"' + return_value + '"}'
            ws.cell(row=i, column=4, value=return_str)

    # 写入 FIELD 和 NOTE
    field_row = lua_name_row + 1 + len(lua_funs)
    note_row = field_row + 1
    value_row = note_row + 1

    # 写入 FIELD 行（绿色）
    field_cell = ws.cell(row=field_row, column=1, value='FIELD')
    field_cell.fill = green_fill
    # 写入 NOTE 行（绿色）
    note_cell = ws.cell(row=note_row, column=1, value='NOTE')
    note_cell.fill = green_fill

    fields = sheet_data.get('fields', [])
    for i, field in enumerate(fields, 2):
        # FIELD 行单元格（绿色）
        field_cell = ws.cell(row=field_row, column=i, value=field.get('FIELD', ''))
        field_cell.fill = green_fill
        # NOTE 行单元格（绿色）
        note_cell = ws.cell(row=note_row, column=i, value=field.get('NOTE', ''))
        note_cell.fill = green_fill

    # 写入 VALUE
    ws.cell(row=value_row, column=1, value='VALUE').fill = yellow_fill

    # 设置样式
    for row in range(1, value_row + 1):
        for column in range(1, len(fields) + 2):
            cell = ws.cell(row=row, column=column)
            cell.font = Font(bold=True)
            # 只有FIELD和NOTE行保持居中对齐，其他行左对齐
            if row == field_row or row == note_row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

def _extract_excel_structure(wb):
    """
    从 Excel 文件中提取结构信息
    :param wb: Workbook 对象
    :return: 结构信息字典，格式与配置字典相同
    """
    import json

    structure = {'sheets': []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            'name': sheet_name,
            'config': {
                'ERL_NAME': '',
                'ERL_INCLUDE': [],
                'ERL_FFUN': [],
                'LUA_NAME': '',
                'LUA_FUN': [],
                'fields': []
            }
        }

        # 提取 ERL_NAME, ERL_INCLUDE, ERL_FUN, LUA_NAME, LUA_FUN
        current_row = 1
        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value

            if cell_value == 'ERL_NAME':
                sheet_data['config']['ERL_NAME'] = ws.cell(row=current_row, column=2).value or ''
            elif cell_value == 'ERL_INCLUDE':
                # 提取所有 ERL_INCLUDE
                column = 2
                while column <= ws.max_column:
                    include = ws.cell(row=current_row, column=column).value
                    if include:
                        sheet_data['config']['ERL_INCLUDE'].append(include)
                    column += 1
            elif cell_value == 'ERL_FUN':
                fun_name = ws.cell(row=current_row, column=2).value or ''
                fun_data = {'name': fun_name, 'params': {'key': [], 'value': []}}

                # 提取 params
                params_str = ws.cell(row=current_row, column=3).value
                if params_str:
                    try:
                        # 修复 JSON 字符串中的单引号问题
                        params_str = params_str.replace("'", '"')
                        # 处理可能的格式问题
                        params_str = params_str.replace('True', 'true').replace('False', 'false').replace('None', 'null')
                        params = json.loads(params_str)
                        fun_data['params']['key'] = params.get('key', [])
                        fun_data['params']['value'] = params.get('value', [])
                    except Exception as e:
                        print(f"解析 ERL_FUN params 失败: {e}")

                # 提取 return
                return_str = ws.cell(row=current_row, column=4).value
                if return_str:
                    try:
                        return_str = return_str.replace("'", '"')
                        return_data = json.loads(return_str)
                        fun_data['return'] = return_data.get('return', '') or ''
                    except Exception as e:
                        print(f"解析 ERL_FUN return 失败: {e}")
                        fun_data['return'] = ''
                else:
                    fun_data['return'] = ''

                # 提取 when
                when_str = ws.cell(row=current_row, column=5).value
                if when_str:
                    try:
                        when_str = when_str.replace("'", '"')
                        when_data = json.loads(when_str)
                        fun_data['when'] = when_data.get('when', '') or ''
                    except Exception as e:
                        print(f"解析 ERL_FUN when 失败: {e}")
                        fun_data['when'] = ''
                else:
                    fun_data['when'] = ''

                # 提取 note
                note_str = ws.cell(row=current_row, column=6).value
                if note_str:
                    # 提取 note 字段，处理 JSON 格式
                    try:
                        note_str = note_str.replace("'", '"')
                        note_data = json.loads(note_str)
                        fun_data['note'] = note_data.get('note', '') or ''
                    except Exception:
                        # 如果不是 JSON 格式，直接使用字符串
                        fun_data['note'] = note_str
                else:
                    fun_data['note'] = ''

                # 提取 fun_note
                fun_note_str = ws.cell(row=current_row, column=7).value
                if fun_note_str:
                    try:
                        fun_note_str = fun_note_str.replace("'", '"')
                        fun_note_data = json.loads(fun_note_str)
                        fun_data['fun_note'] = fun_note_data.get('fun_note', [])
                    except Exception as e:
                        print(f"解析 ERL_FUN fun_note 失败: {e}")

                # 自动生成 fun_note 字段，与 ConfigBuilder 保持一致
                if 'fun_note' not in fun_data:
                    key = fun_data.get('params', {}).get('key', [])
                    value = fun_data.get('params', {}).get('value', [])
                    fun_data['fun_note'] = key + [v for v in value if v not in key]

                sheet_data['config']['ERL_FFUN'].append(fun_data)
            elif cell_value == 'LUA_NAME':
                sheet_data['config']['LUA_NAME'] = ws.cell(row=current_row, column=2).value or ''
            elif cell_value == 'LUA_FUN':
                fun_name = ws.cell(row=current_row, column=2).value or ''
                fun_data = {'name': fun_name, 'params': {'key': [], 'value': []}}

                # 提取 params
                params_str = ws.cell(row=current_row, column=3).value
                if params_str:
                    try:
                        params_str = params_str.replace("'", '"')
                        params_str = params_str.replace('True', 'true').replace('False', 'false').replace('None', 'null')
                        params = json.loads(params_str)
                        fun_data['params']['key'] = params.get('key', [])
                        fun_data['params']['value'] = params.get('value', [])
                    except Exception as e:
                        print(f"解析 LUA_FUN params 失败: {e}")

                # 提取 return
                return_str = ws.cell(row=current_row, column=4).value
                if return_str:
                    try:
                        return_str = return_str.replace("'", '"')
                        return_data = json.loads(return_str)
                        fun_data['return'] = return_data.get('return', '') or ''
                    except Exception as e:
                        print(f"解析 LUA_FUN return 失败: {e}")
                        fun_data['return'] = ''
                else:
                    fun_data['return'] = ''

                sheet_data['config']['LUA_FUN'].append(fun_data)
            elif cell_value == 'FIELD':
                # 提取字段信息
                field_row = current_row
                note_row = field_row + 1

                column = 2
                while column <= ws.max_column:
                    field_name = ws.cell(row=field_row, column=column).value
                    field_note = ws.cell(row=note_row, column=column).value or ''

                    if field_name:
                        sheet_data['config']['fields'].append({
                            'FIELD': field_name,
                            'NOTE': field_note
                        })

                    column += 1

                break

            current_row += 1

        structure['sheets'].append(sheet_data)

    return structure


def _check_structure_diff(config, target_path):
    """
    检查配置与现有文件的结构是否有差异
    :param config: 配置字典
    :param target_path: 目标文件路径
    :return: 是否有差异，True 表示有差异，False 表示无差异
    """
    # 检查目标文件是否存在
    if not os.path.exists(target_path):
        return True

    # 读取现有文件
    try:
        wb = openpyxl.load_workbook(target_path)
    except Exception as e:
        print(f"读取 Excel 文件失败: {e}")
        return True

    # 提取现有文件的结构
    existing_structure = _extract_excel_structure(wb)

    # 只比较 sheets 部分，忽略 filename 字段
    config_sheets = config.get('sheets', [])
    existing_sheets = existing_structure.get('sheets', [])

    # 比较结构
    return existing_sheets != config_sheets

def _write_value_data(ws, value_data):
    """
    写入 VALUE 行数据
    :param ws: Worksheet 对象
    :param value_data: VALUE 行数据，格式为 {field_name: [values], 'value_rows': [row_numbers], 'blank_lines': [row_numbers]}
    """
    # 找到 FIELD 行
    field_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == 'FIELD':
            field_row = row
            break

    if field_row:
        # 提取当前配置的字段顺序
        current_fields = []
        column = 2
        while True:
            field_value = ws.cell(row=field_row, column=column).value
            if field_value:
                current_fields.append(field_value)
                column += 1
            else:
                break

        # 找到所有现有的 VALUE 行
        existing_value_rows = []
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == 'VALUE':
                existing_value_rows.append(row)

        # 写入 VALUE 行数据
        for row_idx, value_row in enumerate(existing_value_rows):
            # 为每一行数据添加 "VALUE" 标识
            ws.cell(row=value_row, column=1, value='VALUE').fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for col_idx, field in enumerate(current_fields):
                if field in value_data and row_idx < len(value_data[field]):
                    value = value_data[field][row_idx]
                else:
                    value = None
                ws.cell(row=value_row, column=col_idx + 2, value=value)

        # 设置样式
        for value_row in existing_value_rows:
            for column in range(1, len(current_fields) + 2):
                cell = ws.cell(row=value_row, column=column)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')