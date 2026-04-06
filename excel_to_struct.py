import os
import json
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook


def _parse_struct_content(content: str) -> Dict[str, Any]:
    """
    解析结构文件内容，提取结构信息
    :param content: 结构文件内容
    :return: 结构信息字典
    """
    import ast

    # 创建一个安全的命名空间
    namespace = {}

    # 执行结构文件内容，获取 config 对象
    try:
        exec(content, namespace)
        config = namespace.get('config', {})
    except Exception as e:
        print(f"解析结构文件内容失败: {e}")
        return {}

    return config


def _check_struct_diff(content: str, struct_file: str) -> bool:
    """
    检查生成的结构与现有结构文件的结构是否有差异
    :param content: 生成的结构文件内容
    :param struct_file: 结构文件路径
    :return: 是否有差异，True 表示有差异，False 表示无差异
    """
    # 检查文件是否存在
    if not os.path.exists(struct_file):
        return True

    # 读取现有文件内容
    try:
        with open(struct_file, 'r', encoding='utf-8') as f:
            existing_content = f.read()
    except Exception as e:
        print(f"读取结构文件失败: {e}")
        return True

    # 解析现有结构
    existing_config = _parse_struct_content(existing_content)

    # 解析当前生成的结构
    current_config = _parse_struct_content(content)

    # 比较结构
    return existing_config != current_config

def excel_to_struct(excel_file: str, struct_file: str) -> None:
    """
    从Excel文件反向生成/更新struct/文件夹中的结构文件

    Args:
        excel_file: Excel文件路径
        struct_file: 生成的结构文件路径
    """
    # 加载Excel文件
    workbook = load_workbook(excel_file)

    # 生成结构文件内容
    content = "from config_builder import ConfigBuilder, SheetBuilder\n\nbuilder = ConfigBuilder()\n\n"

    # 处理每个工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        content += f"### 表格: {sheet_name}\n"
        content += f"sheet = builder.add_sheet(\"{sheet_name}\")\n\n"

        # 解析工作表内容
        erl_name = None
        lua_name = None
        includes = []
        erl_functions = []
        lua_functions = []
        fields = []
        field_names = []
        field_notes = []

        # 第一次遍历：提取字段信息
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not row[0]:
                continue

            config_type = row[0]
            config_value = row[1]

            if config_type == 'FIELD':
                # 提取字段名，只处理非空单元格
                for cell in row[1:]:
                    if cell:
                        field_names.append(cell)
            elif config_type == 'NOTE':
                # 提取字段注释
                for cell in row[1:]:
                    field_notes.append(cell if cell else "")

        # 生成字段代码
        # 确保field_notes的长度与field_names相同
        while len(field_notes) < len(field_names):
            field_notes.append("")
        # 组合字段名和注释
        if field_names:
            content += "# ===field start===\n"
            for field_name, field_note in zip(field_names, field_notes):
                    content += f"sheet.add_field(\"{field_name}\", \"{field_note}\")\n"

        # 第二次遍历：处理其他配置
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not row[0]:
                continue

            config_type = row[0]
            config_value = row[1]

            if config_type == 'ERL_NAME' and config_value:
                erl_name = config_value
                content += "\n# ===erlang start==="
                content += f"\nsheet.set_erl_name(\"{erl_name}\")\n"
            elif config_type == 'LUA_NAME' and config_value:
                lua_name = config_value
                content += "\n# ===lua start==="
                content += f"\nsheet.set_lua_name(\"{lua_name}\")\n"
            elif config_type == 'ERL_INCLUDE' and config_value:
                includes.append(config_value)
                content += f"sheet.add_include(\"{config_value}\")\n"
            elif config_type == 'ERL_FUN' and config_value:
                # 解析ERL_FUN配置
                key = []
                value = []
                if row[2]:  # params
                    try:
                        # 修复JSON字符串中的单引号问题
                        params_str = row[2].replace("'", "\"")
                        params = json.loads(params_str)
                        key = params.get('key', [])
                        value = params.get('value', [])
                    except Exception as e:
                        print(f"解析ERL_FUN params失败: {e}")
                        key = []
                        value = []

                return_type = ""
                if row[3]:  # return
                    try:
                        return_config = json.loads(row[3])
                        return_type = return_config.get('return', "")
                    except Exception as e:
                        print(f"解析ERL_FUN return失败: {e}")
                        pass

                when = ""
                if row[4]:  # when
                    try:
                        when_config = json.loads(row[4])
                        when = when_config.get('when', "")
                    except Exception as e:
                        print(f"解析ERL_FUN when失败: {e}")
                        pass

                note = ""
                if row[5]:  # note
                    note = row[5]

                # 生成ERL函数代码
                content += f"\nsheet.add_erl_function(\n"
                content += f"    name=\"{config_value}\",\n"
                content += f"    key={key},\n"
                content += f"    value={value},\n"
                if return_type:
                    content += f"    return_type=\"{return_type}\",\n"
                if when:
                    content += f"    when=\"{when}\",\n"
                if note:
                    content += f"    note=\"{note}\",\n"
                content += f")\n"
            elif config_type == 'LUA_FUN' and config_value:
                # 解析LUA_FUN配置
                key = []
                value = []
                if row[2]:  # params
                    try:
                        # 修复JSON字符串中的单引号问题
                        params_str = row[2].replace("'", "\"")
                        params = json.loads(params_str)
                        key = params.get('key', [])
                        value = params.get('value', [])
                    except Exception as e:
                        print(f"解析LUA_FUN params失败: {e}")
                        key = []
                        value = []

                return_type = ""
                if row[3]:  # return
                    try:
                        return_config = json.loads(row[3])
                        return_type = return_config.get('return', "")
                    except Exception as e:
                        print(f"解析LUA_FUN return失败: {e}")
                        pass

                # 生成LUA函数代码
                content += f"\nsheet.add_lua_function(\n"
                content += f"    name=\"{config_value}\",\n"
                content += f"    key={key},\n"
                content += f"    value={value},\n"
                if return_type:
                    content += f"    return_type=\"{return_type}\",\n"
                content += f")\n"

        content += "\n"

    # 添加构建配置的代码
    content += "# 构建配置并赋值给全局变量\nconfig = builder.build()"

    # 检查结构是否有差异
    if not _check_struct_diff(content, struct_file):
        print(f"结构文件结构无差异，跳过生成: {struct_file}")
        return

    # 写入文件
    os.makedirs(os.path.dirname(struct_file), exist_ok=True)
    with open(struct_file, 'w', encoding='utf-8') as f:
        f.write(content)

    print(f"成功生成/更新结构文件: {struct_file}")

def process_target_directory() -> None:
    """
    处理target目录下的所有xlsx文件，生成/更新对应的struct文件
    """
    import config as config_module

    # 处理target文件夹路径（支持相对路径和绝对路径）
    if os.path.isabs(config_module.TARGET_FOLDER):
        target_dir = config_module.TARGET_FOLDER
    else:
        target_dir = os.path.join(os.path.dirname(__file__), config_module.TARGET_FOLDER)

    # 处理struct文件夹路径（支持相对路径和绝对路径）
    if os.path.isabs(config_module.STRUCT_FOLDER):
        struct_dir = config_module.STRUCT_FOLDER
    else:
        struct_dir = os.path.join(os.path.dirname(__file__), config_module.STRUCT_FOLDER)

    # 遍历target目录下的所有xlsx文件
    for filename in os.listdir(target_dir):
        if filename.endswith('.xlsx'):
            excel_file = os.path.join(target_dir, filename)
            struct_filename = os.path.splitext(filename)[0] + '.py'
            struct_file = os.path.join(struct_dir, struct_filename)

            print(f"处理文件: {excel_file}")
            excel_to_struct(excel_file, struct_file)

def process_single_excel(excel_name: str) -> None:
    """
    处理单个Excel文件，生成/更新对应的struct文件

    Args:
        excel_name: Excel文件名（不含 .xlsx 后缀）
    """

    import config as config_module

    # 处理target文件夹路径（支持相对路径和绝对路径）
    if os.path.isabs(config_module.TARGET_FOLDER):
        target_dir = config_module.TARGET_FOLDER
    else:
        target_dir = os.path.join(os.path.dirname(__file__), config_module.TARGET_FOLDER)

    # 处理struct文件夹路径（支持相对路径和绝对路径）
    if os.path.isabs(config_module.STRUCT_FOLDER):
        struct_dir = config_module.STRUCT_FOLDER
    else:
        struct_dir = os.path.join(os.path.dirname(__file__), config_module.STRUCT_FOLDER)

    # 构建Excel文件路径和struct文件路径
    excel_file = os.path.join(target_dir, f"{excel_name}.xlsx")
    struct_file = os.path.join(struct_dir, f"{excel_name}.py")

    # 检查Excel文件是否存在
    if not os.path.exists(excel_file):
        print(f"错误: Excel文件 {excel_file} 不存在")
        return

    print(f"处理文件: {excel_file}")
    excel_to_struct(excel_file, struct_file)

if __name__ == "__main__":
    process_target_directory()
